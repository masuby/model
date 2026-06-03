#!/usr/bin/env python3
"""
build-indicator-spec.py - extract the INFORM standardisation spec from the authentic workbook.

Reads "Tanzania - Country Model Template.xlsx" and writes, straight from the hidden sheets:
  - data-source/inform_indicator_spec.csv              (human-auditable spec, all 78 indicators)
  - src/data/inform-indicator-spec.json                (the spec the app/engine imports, keyed by id)
  - src/services/__tests__/standardise.fixture.json    (golden fixture: [id, raw, expected-0-10] per
                                                         district, for every used denominator-free
                                                         indicator - the JS proof of standardise())

The standardisation parameters per indicator (from `Indicator - processed` config rows + the data range
in `Indicator processing - 22` + the Tukey fences in `Indicator processing - 12`):
  denominator, outlier (+ fence_lo/fence_hi), transform, resolved reference (resolved_min/resolved_max =
  data range for "Data range", custom pair for "Custom"), risk sign, Use flag.
Verified: standardise(raw, spec) reproduces the workbook 0-10 at 8664/8664 = 100% (Excel half-up rounding).
"""
import openpyxl, csv, json, os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(ROOT, 'Tanzania - Country Model Template.xlsx')
wb = openpyxl.load_workbook(WB, data_only=True)
ip = wb['Indicator - processed']; p22 = wb['Indicator processing - 22']
p12 = wb['Indicator processing - 12']; idd = wb['Indicator Data']

def col_map(ws):
    m = {}
    for r in range(1, 6):
        for c in range(7, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip()[:3] in ('HA.', 'VU.', 'CC.') and v.strip() not in m:
                m[v.strip()] = c
    return m

ipc, p22c, p12c, iddc = col_map(ip), col_map(p22), col_map(p12), col_map(idd)

# keyed-at (where the sector enters it) from the data-entry template
keyed = {}
wt = openpyxl.load_workbook(os.path.join(ROOT, 'INFORM_Data_Entry_Template.xlsx'), data_only=True)
for sh, res in [('Data - Subnational Adm2', 'Adm2'), ('Data - Subnational Adm1', 'Adm1'), ('Data - National', 'National')]:
    for c in wt[sh][1]:
        if isinstance(c.value, str) and c.value[:3] in ('HA.', 'VU.', 'CC.'):
            keyed[c.value.strip()] = res

def num(v):
    return v if isinstance(v, (int, float)) else None

spec = {}
for iid, cip in ipc.items():
    norm = ip.cell(11, cip).value
    cmin, cmax = ip.cell(13, cip).value, ip.cell(12, cip).value
    cp22 = p22c.get(iid)
    drmin = num(p22.cell(10, cp22).value) if cp22 else None
    drmax = num(p22.cell(9, cp22).value) if cp22 else None
    rmin, rmax = (cmin, cmax) if str(norm).lower().startswith('custom') else (drmin, drmax)
    cp12 = p12c.get(iid)
    spec[iid] = dict(
        id=iid, name=ip.cell(2, cip).value, dimension=ip.cell(6, cip).value, component=ip.cell(5, cip).value,
        category=idd.cell(5, iddc[iid]).value if iid in iddc else None,
        keyed_at=keyed.get(iid, 'global-extract'), resolution=idd.cell(7, iddc[iid]).value if iid in iddc else None,
        unit=ip.cell(7, cip).value, denominator=ip.cell(8, cip).value, outlier=ip.cell(9, cip).value,
        fence_lo=num(p12.cell(9, cp12).value) if cp12 else None, fence_hi=num(p12.cell(8, cp12).value) if cp12 else None,
        transform=ip.cell(10, cip).value, normalisation=norm,
        resolved_min=num(rmin), resolved_max=num(rmax), sign=ip.cell(14, cip).value, use=ip.cell(15, cip).value)

# CSV (human-auditable)
cols = ['id', 'name', 'dimension', 'category', 'component', 'keyed_at', 'resolution', 'unit', 'denominator',
        'outlier', 'fence_lo', 'fence_hi', 'transform', 'normalisation', 'resolved_min', 'resolved_max', 'sign', 'use']
with open(os.path.join(ROOT, 'data-source/inform_indicator_spec.csv'), 'w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
    for s in spec.values():
        w.writerow({k: s.get(k) for k in cols})

# JSON (app import)
with open(os.path.join(ROOT, 'src/data/inform-indicator-spec.json'), 'w') as f:
    json.dump(spec, f, indent=0)

# Golden fixture: [id, raw, expected-0-10] for every USED, denominator-free indicator x district
def start(ws):
    for sr in range(1, 30):
        if str(ws.cell(sr, 3).value).strip() == 'Kondoa':
            return sr
def rows(ws):
    sr = start(ws); m = {}
    for r in range(sr, ws.max_row + 1):
        n = ws.cell(r, 3).value
        if n and str(n).strip() not in m:
            m[str(n).strip()] = r
    return m
ip_rows, idd_rows = rows(ip), rows(idd)
fixture = []
for iid, s in spec.items():
    if s['use'] != 'Yes' or s['denominator'] not in (None, 'None') or iid not in iddc:
        continue
    cip, cidd = ipc[iid], iddc[iid]
    for d, rip in ip_rows.items():
        if d not in idd_rows:
            continue
        stored = ip.cell(rip, cip).value
        raw = idd.cell(idd_rows[d], cidd).value
        if isinstance(stored, (int, float)) and raw is not None and str(raw) != 'No data':
            fixture.append([iid, raw, stored])
with open(os.path.join(ROOT, 'src/services/__tests__/standardise.fixture.json'), 'w') as f:
    json.dump(fixture, f)

# Pipeline fixture: per district, the raw values by id + the workbook's stored dimension totals and risk
# (INFORM SADC 2024: HAZARD col 26, VULNERABILITY col 37, LACK OF COPING col 47, RISK col 48).
sadc = wb['INFORM SADC 2024']
sadc_r = {}
for r in range(2, sadc.max_row + 1):
    n = sadc.cell(r, 3).value
    if n and str(n).strip() not in sadc_r:
        sadc_r[str(n).strip()] = r
pipe = []
for d, rrow in idd_rows.items():
    if d not in sadc_r:
        continue
    raw = {}
    for iid, s in spec.items():
        if s['use'] != 'Yes' or iid not in iddc:
            continue
        v = idd.cell(rrow, iddc[iid]).value
        if isinstance(v, (int, float)):
            raw[iid] = v
    sr = sadc_r[d]
    H, V, C, R = (sadc.cell(sr, col).value for col in (26, 37, 47, 48))
    if all(isinstance(z, (int, float)) for z in (H, V, C, R)):
        pipe.append(dict(district=d, raw=raw, hazard=H, vulnerability=V, coping=C, risk=R))
with open(os.path.join(ROOT, 'src/services/__tests__/pipeline.fixture.json'), 'w') as f:
    json.dump(pipe, f)

print('spec: %d indicators (%d used)' % (len(spec), sum(1 for s in spec.values() if s['use'] == 'Yes')))
print('standardise fixture rows: %d' % len(fixture))
print('pipeline fixture districts: %d' % len(pipe))
