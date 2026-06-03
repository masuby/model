#!/usr/bin/env python3
"""
build-collection-tool.py - generate the ONE shared INFORM Tanzania data-collection workbook (NORMAL / v1).

Sent to all sectors. Each finds its rows (filter the Sector column), fills the ACTUAL VALUE in the
indicator's own unit, and the LOCKED standardisation formula shows the 0-10 live - exactly the proven
standardise() (denominator None here, optional log, min-max vs the FROZEN reference, Decrease inversion,
clamp, Excel ROUND). Unfilled rows are fine (the engine uses the foundation / skips them). Only the
ACTUAL VALUE column is unlocked; everything else is protected so the standardiser stays consistent.

The ADVANCED version (v2) is the same tool with extra sector sub-indicator rows (SPEI, soil moisture,
water levels, biomass) in a research tab - built later, on top of this.

Output: public/INFORM_TZ_Collection_Tool.xlsx  (downloadable from the site to send to sectors).
"""
import json, os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.protection import SheetProtection

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SPEC = json.load(open(os.path.join(ROOT, 'src/data/inform-indicator-spec.json')))

# responsible Tanzanian sector per component (who fills it locally). Global-foundation indicators are
# pre-filled by the centre; here we tag the sector that would provide LOCAL data.
def norm(s): return ''.join(ch for ch in str(s).lower() if ch.isalpha())
# ONE key Tanzanian stakeholder per indicator - the institution that OWNS the data (deep institutional
# assessment; no shared slashes). Weather/climate=TMA, geology=GST, environment=NEMC, forests=TFS,
# agriculture/food=MoA, security/refugees=MoHA, police=TPF, statistics=NBS, finance=MoFP, health=MoH,
# water=MoW, telecom=TCRA, education=MoEST, disaster-mgmt=PMO-DMD, local-government=PO-RALG.
SECTOR = {
    'drought': 'TMA', 'flood': 'TMA', 'stormscyclone': 'TMA', 'lightning': 'TMA', 'heatwave': 'TMA',
    'earthquake': 'GST', 'landslide': 'GST', 'volcano': 'GST',
    'environmentaldegradation': 'NEMC', 'soilerosion': 'NEMC', 'coastalhazards': 'NEMC', 'hazardousmaterial': 'NEMC',
    'wildfire': 'TFS', 'zoonosesplantspests': 'MoA',
    'conflictrisk': 'MoHA', 'conflictintensity': 'MoHA', 'internalviolence': 'TPF', 'vehicleaccidents': 'TPF',
    'developmentpoverty': 'NBS', 'economicdependency': 'MoFP', 'habitat': 'NBS', 'livelihoods': 'MoA',
    'displacedpeople': 'MoHA', 'healthconditions': 'MoH', 'childrenhealthandnutrition': 'MoH', 'economic': 'NBS',
    'accesstohealthcare': 'MoH', 'economiccapacity': 'NBS', 'wash': 'MoW', 'communication': 'TCRA',
    'education': 'MoEST', 'drrimplementation': 'PMO-DMD', 'governance': 'PO-RALG',
}
def sector_of(component): return SECTOR.get(norm(component), 'INFORM (global)')

# Tool rows: ALL used indicators - exactly the regional INFORM set, UNTAMPERED. The formula reproduces the
# hidden sheets faithfully (cap -> transform None/Log/Exp -> min-max -> sign -> clamp -> ROUND), wrapped in
# IFERROR -> "No data" exactly as the workbook, so degenerate cases (health-facility density custom[0,0])
# show "No data" like INFORM rather than being dropped.
rows = [s for s in SPEC.values() if s['use'] == 'Yes']
order = {'Hazards & Exposure': 0, 'Vulnerability': 1, 'Coping Capacity': 2}
rows.sort(key=lambda s: (order.get(s['dimension'], 9), s['category'], s['component'], s['name']))

wb = openpyxl.Workbook()

# ---- Sheet 1: How to use ----
how = wb.active; how.title = 'How to use'
how.sheet_view.showGridLines = False
intro = [
    ('INFORM Tanzania - Data Collection Tool (NORMAL / v1)', 14, True),
    ('', 10, False),
    ('1.  Filter the "Sector" column on the Data Entry sheet to find YOUR indicators.', 11, False),
    ('2.  For each, type the ACTUAL VALUE in its own unit (e.g. 18 for 18% underweight, 11 for an', 11, False),
    ('     11-year drought frequency, a count of displaced people). The unit is shown on each row.', 11, False),
    ('3.  The "0-10 (auto)" column standardises it instantly - exactly as the official INFORM workbook.', 11, False),
    ('     You never invent a score; the tool computes it. These cells are locked.', 11, False),
    ('4.  Leave rows you do not have data for BLANK - the model uses the foundation value, nothing breaks.', 11, False),
    ('5.  Save and send the file back. The centre reads only the 0-10 values; your raw stays with you.', 11, False),
    ('', 10, False),
    ('Reference = the fixed Min/Max the value is scaled against (frozen, so a value means the same at', 10, False),
    ('district, region, country, or village). Direction "protective" means higher is better (inverted).', 10, False),
    ('Only the ACTUAL VALUE column is editable. The ADVANCED version adds research rows (SPEI, soil', 10, False),
    ('moisture, water levels...) for sectors to refine an indicator beyond the legacy single source.', 10, False),
]
for i, (txt, sz, bold) in enumerate(intro, 1):
    c = how.cell(i, 1, txt); c.font = Font(size=sz, bold=bold, color='1F3A5F' if bold else '000000')
how.column_dimensions['A'].width = 100

# ---- Data Entry (per-indicator) - kept as a HIDDEN reference, superseded by the 195-Council capture ----
ws = wb.create_sheet('Data Entry')
ws.sheet_view.showGridLines = False
ws.sheet_state = 'hidden'
headers = ['Sector', 'Dimension', 'Component', 'Indicator', 'Unit', 'Level', 'Transform',
           'Ref Min', 'Ref Max', 'Direction', 'ACTUAL VALUE', '0-10 (auto)']
HEAD = PatternFill('solid', fgColor='1F3A5F'); thin = Side(style='thin', color='D0D7E2')
for j, h in enumerate(headers, 1):
    c = ws.cell(1, j, h); c.font = Font(bold=True, color='FFFFFF', size=10)
    c.fill = HEAD; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

INPUT = PatternFill('solid', fgColor='FFF7E6'); AUTO = PatternFill('solid', fgColor='EAF3EA')
for i, s in enumerate(rows, start=2):
    direction = 'protective' if str(s['sign']).startswith('Dec') else 'risk'
    vals = [sector_of(s['component']), s['dimension'], s['category'] + ' / ' + s['component'], s['name'],
            s.get('unit') or 'value', s.get('keyed_at') or '', s.get('transform') or 'None',
            s.get('resolved_min'), s.get('resolved_max'), direction]
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v); c.font = Font(size=10); c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # ACTUAL VALUE (unlocked input)
    cin = ws.cell(i, 11); cin.fill = INPUT; cin.protection = openpyxl.styles.Protection(locked=False)
    cin.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # 0-10 (locked formula) - the standardiser, in Excel, referencing this row's ref/transform/direction
    g, h, k, sgn = f'G{i}', f'H{i}', f'I{i}', f'J{i}'  # transform, min, max, direction
    av = f'K{i}'
    # outlier cap (Tukey fence), embedded per row for the few indicators with outlier detection on
    if s.get('outlier') == 'Yes' and isinstance(s.get('fence_lo'), (int, float)) and isinstance(s.get('fence_hi'), (int, float)):
        av = f'MAX(MIN(K{i},{s["fence_hi"]}),{s["fence_lo"]})'
    x = f'IF({g}="Logarithm",LN(0.001+{av}),IF({g}="Exponential",EXP({av}),{av}))'  # transform None/Log/Exp
    base = f'10*(({x})-{h})/({k}-{h})'                   # min-max
    expr = f'IF({sgn}="protective",10-{base},{base})'    # Decrease inversion
    formula = f'=IF(K{i}="","",IFERROR(MAX(0,MIN(10,ROUND({expr},1))),"No data"))'
    cf = ws.cell(i, 12, formula); cf.fill = AUTO; cf.font = Font(size=10, bold=True, color='1F6F3D')
    cf.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:L{len(rows) + 1}'
widths = [16, 18, 30, 30, 12, 10, 11, 9, 9, 11, 14, 11]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
# protect everything except the ACTUAL VALUE column
ws.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False)

# ---- Sheets 3 + 4: 195-COUNCIL capture (production resolution) ----
# Council Entry = the 195 NBS councils (rows) x indicators (cols), one unlocked value per cell.
# Council 0-10  = the same grid with the VALIDATED locked standardiser formula per cell. Unfilled cells
# stay blank (the model uses the foundation - bottom-up). Formula identical to the proven Data Entry sheet.
from openpyxl.utils import get_column_letter
geo = json.load(open(os.path.join(ROOT, 'src/data/tanzania-councils.json')))
councils = sorted({(str(f['properties'].get('reg') or f['properties'].get('adm1Name') or ''),
                    str(f['properties'].get('name') or ''), str(f['properties'].get('code') or ''))
                   for f in geo['features']})
entry = wb.create_sheet('Council Entry'); sc = wb.create_sheet('Council 0-10')
entry.sheet_view.showGridLines = False; sc.sheet_view.showGridLines = False
meta_rows = ['Indicator', 'Unit', 'Sector', 'Transform', 'Ref Min', 'Ref Max', 'Direction', 'Fence Lo', 'Fence Hi']
HEADR = 10  # column-title row; councils start at HEADR+1
for r, lbl in enumerate(meta_rows, 1):
    entry.cell(r, 3, lbl).font = Font(size=9, italic=True, color='64748B')
for j, ttl in enumerate(['Region', 'Council', 'Code'], 1):
    for sh in (entry, sc):
        c = sh.cell(HEADR, j, ttl); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = HEAD
for ci, s in enumerate(rows):
    col = 4 + ci; L = get_column_letter(col)
    sgn = 'Decrease Risk' if str(s['sign']).startswith('Dec') else 'Increase Risk'
    flo = s.get('fence_lo') if s.get('outlier') == 'Yes' else None
    fhi = s.get('fence_hi') if s.get('outlier') == 'Yes' else None
    for r, v in enumerate([s['name'], s.get('unit') or 'value', sector_of(s['component']), s.get('transform') or 'None',
                           s.get('resolved_min'), s.get('resolved_max'), sgn, flo, fhi], 1):
        entry.cell(r, col, v).font = Font(size=9)
    for sh in (entry, sc):
        c = sh.cell(HEADR, col, s['name']); c.font = Font(bold=True, color='FFFFFF', size=9); c.fill = HEAD
        sh.column_dimensions[L].width = 13
for ri, (reg, name, code) in enumerate(councils):
    r = HEADR + 1 + ri
    for sh in (entry, sc):
        sh.cell(r, 1, reg); sh.cell(r, 2, name); sh.cell(r, 3, code)
    for ci, s in enumerate(rows):
        col = 4 + ci; L = get_column_letter(col)
        ec = entry.cell(r, col); ec.fill = INPUT; ec.protection = openpyxl.styles.Protection(locked=False)
        val = f"'Council Entry'!{L}{r}"; tf = f"'Council Entry'!{L}$4"; mn = f"'Council Entry'!{L}$5"
        mx = f"'Council Entry'!{L}$6"; sg = f"'Council Entry'!{L}$7"; flc = f"'Council Entry'!{L}$8"; fhc = f"'Council Entry'!{L}$9"
        cap = f"IF(AND(ISNUMBER({flc}),ISNUMBER({fhc})),MAX(MIN({val},{fhc}),{flc}),{val})"
        x = f'IF({tf}="Logarithm",LN(0.001+{cap}),IF({tf}="Exponential",EXP({cap}),{cap}))'
        base = f"10*(({x})-{mn})/({mx}-{mn})"
        expr = f'IF({sg}="Decrease Risk",10-{base},{base})'
        sc.cell(r, col, f'=IF({val}="","",IFERROR(MAX(0,MIN(10,ROUND({expr},1))),"No data"))').font = Font(size=9, color='1F6F3D')
for sh in (entry, sc):
    sh.freeze_panes = 'D11'
    sh.column_dimensions['A'].width = 16; sh.column_dimensions['B'].width = 22; sh.column_dimensions['C'].width = 8
# protect the sheet but ALLOW selecting/editing unlocked (yellow) cells - selectUnlockedCells=False
entry.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False)
sc.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False)

# ---- Sheet 5: ADVANCED (research) - v2, fully UNLOCKED ----
# The locked NORMAL/Council sheets above are the regional INFORM, untouched. This tab lets sectors RESEARCH
# advancements: extra sub-indicators (SPEI, soil moisture, MoW water levels) and INFORM gaps (heatwave,
# lightning), with literature-cited references + proposed weights. Same standardiser engine; everything editable.
adv = list(csv.DictReader(open(os.path.join(ROOT, 'data-source/inform_advanced_spec.csv'))))
aw = wb.create_sheet('Advanced (research)'); aw.sheet_view.showGridLines = False
aw.sheet_state = 'hidden'  # NOT shared/seen by sectors; kept as a complete sheet "down" in the file for us
aw.cell(1, 1, 'ADVANCED (research) - v2 PROPOSAL.  The locked NORMAL + Council sheets are the regional INFORM, unchanged.').font = Font(bold=True, size=11, color='B45309')
aw.cell(2, 1, 'Refine an indicator with extra sub-indicators, or fill an INFORM gap. References are literature-cited and PROPOSED (so are the weights); when validated they get enforced. Everything on this tab is editable.').font = Font(size=9, italic=True, color='64748B')
AHEAD = ['Component', 'Sector', 'Indicator', 'Unit', 'Transform', 'Ref Min', 'Ref Max', 'Direction', 'Weight', 'Basis (literature)', 'ACTUAL VALUE', '0-10 (auto)']
ADVFILL = PatternFill('solid', fgColor='B45309')
hr = 4
for j, h in enumerate(AHEAD, 1):
    c = aw.cell(hr, j, h); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = ADVFILL; c.alignment = Alignment(horizontal='center', wrap_text=True)
for i, a in enumerate(adv):
    r = hr + 1 + i
    direction = 'protective' if str(a['sign']).startswith('Dec') else 'risk'
    for j, v in enumerate([a['component'], a['sector'], a['name'], a['unit'], a['transform'],
                           float(a['resolved_min']), float(a['resolved_max']), direction, float(a['weight']), a['basis']], 1):
        aw.cell(r, j, v).font = Font(size=9)
    aw.cell(r, 11).fill = INPUT  # ACTUAL VALUE
    g, mn, mx, sg, av = f'E{r}', f'F{r}', f'G{r}', f'H{r}', f'K{r}'
    x = f'IF({g}="Logarithm",LN(0.001+{av}),IF({g}="Exponential",EXP({av}),{av}))'
    base = f'10*(({x})-{mn})/({mx}-{mn})'
    expr = f'IF({sg}="protective",10-{base},{base})'
    aw.cell(r, 12, f'=IF({av}="","",IFERROR(MAX(0,MIN(10,ROUND({expr},1))),"No data"))').font = Font(size=9, bold=True, color='1F6F3D')
aw.freeze_panes = 'A5'
for j, w in enumerate([16, 14, 28, 12, 10, 9, 9, 11, 8, 42, 14, 11], 1):
    aw.column_dimensions[get_column_letter(j)].width = w
# (no SheetProtection - fully unlocked for research)

out = os.path.join(ROOT, 'public/INFORM_TZ_Collection_Tool.xlsx')
wb.save(out)
print('council rows:', len(councils), '| advanced research rows:', len(adv))
print('wrote', out)
print('indicators (rows):', len(rows))
from collections import Counter
print('by sector:')
for k, v in Counter(sector_of(s['component']) for s in rows).most_common():
    print('  %-16s %d' % (k, v))
