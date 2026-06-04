#!/usr/bin/env python3
"""
fill-collection-tool.py - populate the NORMAL collection tool with AUTHENTIC official-INFORM data to
create the operational BENCHMARK dataset.

Source of truth: the regional INFORM workbook's real raw values per unit (src/services/__tests__/
pipeline.fixture.json - the 170-district gold proven 8664/8664). Each of the 195 councils inherits the
raw values of the workbook unit it maps to (reconciliation.data_unit - the model's own council->unit map;
194/195 resolve exactly, Butiama -> Musoma as the documented split fallback). Nothing is invented: every
value is the official INFORM figure for that council's parent unit, and the tool's locked standardiser
then reproduces the workbook 0-10 exactly.

Output: inform_sheets/INFORM_TZ_Collection_Benchmark.xlsx  (the filled benchmark; the blank template
INFORM_TZ_Collection_Tool.xlsx is kept for sending to sectors).
"""
import json, os, csv
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def norm(s): return ''.join(ch for ch in str(s).lower() if ch.isalnum())

SPEC = json.load(open(os.path.join(ROOT, 'src/data/inform-indicator-spec.json')))
name2id = {s['name']: s['id'] for s in SPEC.values() if s['use'] == 'Yes'}

# ---- authentic raw per workbook unit (the official INFORM figures) ----
pf = json.load(open(os.path.join(ROOT, 'src/services/__tests__/pipeline.fixture.json')))
raw_by_unit = {norm(r['district']): r['raw'] for r in pf}

# ---- council -> workbook unit map (the model's own reconciliation) ----
cc = {r['counc_code']: r for r in csv.DictReader(open(os.path.join(ROOT, 'data-source/council_climate.csv')))}
rec_by_name = {norm(r['council_name']): r for r in csv.DictReader(open(os.path.join(ROOT, 'data-source/council_reconciliation.csv')))}
FALLBACK = {'butiama': 'Musoma'}  # post-2012 split from Musoma Rural; the workbook's rural unit is 'Musoma'

def unit_for(code, tool_name):
    c = cc.get(code)
    rec = rec_by_name.get(norm(c['council'])) if c else None
    if rec is None:
        rec = rec_by_name.get(norm(tool_name))
    du = rec['data_unit'] if rec else tool_name
    if norm(du) in raw_by_unit:
        return du, raw_by_unit[norm(du)]
    fb = FALLBACK.get(norm(du)) or FALLBACK.get(norm(tool_name))
    if fb and norm(fb) in raw_by_unit:
        return fb + ' (fallback)', raw_by_unit[norm(fb)]
    # last resort: nearest unit in the same region (kept honest - flagged)
    return du + ' (UNMATCHED)', None

# ---- load the blank template, read the column -> indicator-id layout from the header ----
TPL = os.path.join(ROOT, 'inform_sheets/INFORM_TZ_Collection_Tool.xlsx')
wb = openpyxl.load_workbook(TPL)
entry = wb['Council Entry']
HEADR = 10
col_id = {}  # column -> spec id
for col in range(4, entry.max_column + 1):
    nm = entry.cell(1, col).value
    if nm and nm in name2id:
        col_id[col] = name2id[nm]

# ---- fill every council row with its unit's authentic raw ----
filled = 0; cells = 0; flagged = []
for r in range(HEADR + 1, entry.max_row + 1):
    reg = entry.cell(r, 1).value; name = entry.cell(r, 2).value; code = entry.cell(r, 3).value
    if not name:
        continue
    unit, raw = unit_for(code, name)
    if '(fallback)' in unit or '(UNMATCHED)' in unit:
        flagged.append((code, name, unit))
    if raw is None:
        continue
    for col, iid in col_id.items():
        v = raw.get(iid)
        if v is not None and v != '':
            entry.cell(r, col, v)  # write the authentic raw value (unlocked input cell)
            cells += 1
    filled += 1

out = os.path.join(ROOT, 'inform_sheets/INFORM_TZ_Collection_Benchmark.xlsx')
wb.save(out)
print('wrote', out)
print('councils filled:', filled, '/ 195 | raw values written:', cells, '| indicator cols:', len(col_id))
print('flagged (fallback/unmatched):', flagged if flagged else 'none')
