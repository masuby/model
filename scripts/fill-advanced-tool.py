#!/usr/bin/env python3
"""
fill-advanced-tool.py - populate the ADVANCED tool with AUTHENTIC on-disk EO computations to create the
advanced BENCHMARK. Only sub-indicators we genuinely have data for are filled; every other source is left
BLANK (the no-bias design skips it). Nothing is invented.

Authentic on-disk sources -> advanced sub-indicators:
  DR-SPEI  <- chirps_drought_spi_spei.spei_severity   (CHIRPS v3 / ERA5, per district)
  DR-CV    <- chirps_drought_spi_spei.rain_cv          (CHIRPS v3 interannual CV)
  DR-SEAS  <- chirps_drought_spi_spei.season_fail_freq (growing-season failure)
  FL-R50   <- council_climate.heavy_days_yr            (CHIRPS v3 daily >50 mm, per council)
  FL-R100  <- council_climate.vheavy_days_yr           (CHIRPS v3 daily >100 mm, per council)
  EQ-HIST  <- earthquake_usgs.n_events_50km            (USGS catalogue, historical seismicity density)
Left blank (genuine gaps - see docs/INDICATOR_SOURCE_REGISTER.md for the authentic upgrade path):
  aridity, NDVI, soil moisture, water levels, river peak, basin records, PGA, susceptibility,
  storms/cyclone (coastal hotspot score, not an events/decade rate), coastal waves/wind/surge/erosion/
  mangrove, heatwave-days, lightning.

Output: inform_sheets/INFORM_TZ_Advanced_Benchmark.xlsx
"""
import json, os, csv
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def norm(s): return ''.join(ch for ch in str(s).lower() if ch.isalnum())

# advanced sub-indicator name -> id
adv = list(csv.DictReader(open(os.path.join(ROOT, 'data-source/inform_advanced_spec.csv'))))
name2id = {a['name']: a['id'] for a in adv}

# ---- council -> workbook unit (district) map, same chain as the normal fill ----
cc = {r['counc_code']: r for r in csv.DictReader(open(os.path.join(ROOT, 'data-source/council_climate.csv')))}
rec_by_name = {norm(r['council_name']): r for r in csv.DictReader(open(os.path.join(ROOT, 'data-source/council_reconciliation.csv')))}
FALLBACK = {'butiama': 'Musoma'}
def data_unit(code, tool_name):
    c = cc.get(code)
    rec = rec_by_name.get(norm(c['council'])) if c else None
    if rec is None: rec = rec_by_name.get(norm(tool_name))
    du = rec['data_unit'] if rec else tool_name
    return FALLBACK.get(norm(du), du)

# ---- authentic sources ----
drought = {norm(r['dist_name']): r for r in csv.DictReader(open(os.path.join(ROOT, 'data-source/chirps_drought_spi_spei.csv')))}
quake = {norm(r['dist_name']): r for r in csv.DictReader(open(os.path.join(ROOT, 'data-source/earthquake_usgs.csv')))}
# aridity computed per council (CHIRPS v3 + ERA5 t2m Thornthwaite); keyed by council code
ARIDF = os.path.join(ROOT, 'data-source/aridity_councils.csv')
aridity = {r['code']: r['aridity_index'] for r in csv.DictReader(open(ARIDF))} if os.path.exists(ARIDF) else {}
# GEE EO per council (NDVI greenness, SMAP soil moisture, Sentinel-1 flood frequency); keyed by code
def _load(fn, col):
    p = os.path.join(ROOT, 'data-source', fn)
    return {r['code']: r[col] for r in csv.DictReader(open(p))} if os.path.exists(p) else {}
ndvi_c = _load('ndvi_councils.csv', 'ndvi')
soil_c = _load('soil_moisture_councils.csv', 'ssm')
flood_c = _load('flood_s1_councils.csv', 'floodfreq')

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

# value resolver per (council code, tool_name) -> {adv_id: value}
def values_for(code, tool_name):
    out = {}
    c = cc.get(code)
    if c:  # per-council CHIRPS daily heavy-rain counts
        out[name2id['Extreme wet days (>50 mm)']] = num(c.get('heavy_days_yr'))
        out[name2id['Very-heavy days (>100 mm)']] = num(c.get('vheavy_days_yr'))
    if code in aridity:  # per-council aridity P/PET (CHIRPS v3 + ERA5 Thornthwaite)
        out[name2id['Aridity (P/PET)']] = num(aridity.get(code))
    if code in ndvi_c:   # GEE EO (MODIS NDVI, SMAP soil moisture, Sentinel-1 flood)
        out[name2id['Mean NDVI (greenness)']] = num(ndvi_c.get(code))
    if code in soil_c:
        out[name2id['Soil moisture']] = num(soil_c.get(code))
    # NOTE: Sentinel-1 flood (flood_c) intentionally NOT wired in - the simple VV<-17dB threshold
    # failed validation (flags soda-lakes/salt-flats e.g. Ngorongoro/Lake Natron as 'water', not floods).
    # Kept as a flagged research artifact (flood_s1_councils.csv); needs the Paper-1 change-detection method.
    du = norm(data_unit(code, tool_name))
    d = drought.get(du)
    if d:
        out[name2id['SPEI-12 drought depth']] = num(d.get('spei_severity'))
        out[name2id['Rainfall variability (CV)']] = num(d.get('rain_cv'))
        out[name2id['Growing-season failure frequency']] = num(d.get('season_fail_freq'))
    q = quake.get(du)
    if q:
        out[name2id['Historical seismicity density']] = num(q.get('n_events_50km'))
    return {k: v for k, v in out.items() if v is not None}

# ---- load template, map column -> advanced id (header row 1 = name) ----
TPL = os.path.join(ROOT, 'inform_sheets/INFORM_TZ_Advanced_Tool.xlsx')
wb = openpyxl.load_workbook(TPL)
e = wb['Advanced Entry']
HEADR = 10
col_id = {c: name2id[e.cell(1, c).value] for c in range(4, e.max_column + 1) if e.cell(1, c).value in name2id}

filled = cells = 0; bycount = {}
for r in range(HEADR + 1, e.max_row + 1):
    name = e.cell(r, 2).value; code = e.cell(r, 3).value
    if not name: continue
    vals = values_for(code, name)
    if vals: filled += 1
    for col, iid in col_id.items():
        if iid in vals:
            e.cell(r, col, vals[iid]); cells += 1
            bycount[iid] = bycount.get(iid, 0) + 1

out = os.path.join(ROOT, 'inform_sheets/INFORM_TZ_Advanced_Benchmark.xlsx')
wb.save(out)
print('wrote', out)
print('councils with >=1 authentic value:', filled, '/ 195 | cells written:', cells)
id2name = {v: k for k, v in name2id.items()}
for iid, n in sorted(bycount.items()):
    print('  %-18s %-34s %d councils' % (iid, id2name[iid][:34], n))
