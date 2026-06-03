#!/usr/bin/env python3
"""
build-advanced-tool.py - the ADVANCED (v2) workbook: a COMPLETELY SEPARATE Excel that EXPLODES each
hazard into ALL the institutions that hold relevant evidence, each with its OWN cell.

Idea (the user's): "exploding the model, capturing the issue without biasing". A component is no longer
one number from one owner; it is a BASKET of sub-indicators, each owned by the institution that has the
data - and any institution can contribute to any component:
  - Drought   = TMA (SPEI, aridity, rainfall variability) + MoA (season-failure, NDVI, soil moisture) + MoW (water levels)
  - Flood     = TMA (>50/>100 mm days) + MoW (river peak) + BWB (basin flood records) + PMO-DMD (recorded events)
  - Landslide = GST (susceptibility hotspots) + TMA (triggering rainfall)
  - Coastal   = TMA (waves, onshore wind, storm surge) + NEMC (shoreline erosion, mangrove loss)   <- cross-cutting
  - Storms / Earthquake / Heatwave / Lightning likewise.

The lead institution (the NORMAL tool's single owner) stays accountable for the component; here it is
JOINED by everyone else who has evidence. Each contributor fills its own cell in its own unit; the LOCKED
standardiser shows 0-10 (the proven INFORM pipeline). The component then aggregates by WEIGHTED AVERAGE
**over the cells that are present** - a BLANK is SKIPPED, it does NOT bias the result. So if in council X
only TMA rainfall is available for landslide and GST susceptibility is blank, the component is simply the
TMA evidence; it does not become "the case" for everyone, and a missing source never drags the number down.

Output: public/INFORM_TZ_Advanced_Tool.xlsx  (a separate download, NOT shared with the normal tool).
"""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
adv = list(csv.DictReader(open(os.path.join(ROOT, 'data-source/inform_advanced_spec.csv'))))

# group by component, keep a stable component order (as first seen), so each component's sub-indicators
# occupy a CONTIGUOUS block of columns (needed for the present-only weighted-average SUMPRODUCT range).
comp_order = []
for a in adv:
    if a['component'] not in comp_order:
        comp_order.append(a['component'])
adv.sort(key=lambda a: (comp_order.index(a['component']),))

# one colour per institution so "who fills what" is obvious at a glance
INST_FILL = {
    'TMA':     'E3F0FB',  # sky
    'MoA':     'E8F4E5',  # green
    'MoW':     'E0F2F1',  # teal
    'BWB':     'E0F7FA',  # cyan
    'PMO-DMD': 'FFF3E0',  # amber
    'GST':     'F3ECE2',  # earth
    'NEMC':    'EDE7F6',  # violet
}
def inst_fill(sec): return PatternFill('solid', fgColor=INST_FILL.get(sec, 'FFF7E6'))

geo = __import__('json').load(open(os.path.join(ROOT, 'src/data/tanzania-councils.json')))
councils = sorted({(str(f['properties'].get('reg') or f['properties'].get('adm1Name') or ''),
                    str(f['properties'].get('name') or ''), str(f['properties'].get('code') or ''))
                   for f in geo['features']})

wb = openpyxl.Workbook()
HEAD = PatternFill('solid', fgColor='B45309'); thin = Side(style='thin', color='E5DACb'.upper())
AUTO = PatternFill('solid', fgColor='EAF3EA')

# ---- Sheet 1: How to use ----
how = wb.active; how.title = 'How to use'; how.sheet_view.showGridLines = False
intro = [
    ('INFORM Tanzania - ADVANCED Data Collection Tool (v2, EXPLODED multi-source)', 14, True),
    ('', 10, False),
    ('This workbook is SEPARATE from the shared NORMAL tool. Here each hazard is EXPLODED into every', 11, False),
    ('institution that holds relevant evidence - each with its OWN cell, in its own unit.', 11, False),
    ('', 8, False),
    ('  Drought    = TMA (SPEI, aridity, rainfall variability) + MoA (season-failure, NDVI, soil) + MoW (water levels)', 10, False),
    ('  Flood      = TMA (>50 / >100 mm days) + MoW (river peak) + BWB (basin flood records) + PMO-DMD (events)', 10, False),
    ('  Landslide  = GST (susceptibility hotspots) + TMA (triggering rainfall)', 10, False),
    ('  Coastal    = TMA (waves, onshore wind, storm surge) + NEMC (shoreline erosion, mangrove loss)', 10, False),
    ('  Storms / Earthquake / Heatwave / Lightning likewise.', 10, False),
    ('', 8, False),
    ('1.  Find YOUR institution\'s coloured columns on the "Advanced Entry" sheet (the Owner row names them).', 11, False),
    ('2.  Type the ACTUAL VALUE where you have it. Leave BLANK where you do not - a blank is SKIPPED.', 11, False),
    ('3.  "Advanced 0-10" standardises each cell (the proven INFORM pipeline, locked).', 11, False),
    ('4.  "Components" aggregates each basket by WEIGHTED AVERAGE over the cells PRESENT - so a missing', 11, False),
    ('     source never biases the result. If only TMA rainfall is present for a council\'s landslide,', 11, False),
    ('     the component is just that evidence; it does not "become the case" for everyone else.', 11, False),
    ('', 8, False),
    ('Weights are research-PROPOSED (literature-cited in the Basis column) and will be enforced once', 10, False),
    ('validated. Nothing here replaces the regional INFORM - it refines it where local data exists.', 10, False),
]
for i, (txt, sz, bold) in enumerate(intro, 1):
    how.cell(i, 1, txt).font = Font(size=sz, bold=bold, color='B45309' if bold else '000000')
how.column_dimensions['A'].width = 112

# ---- Sheet 2: Advanced Entry  (195 councils x exploded sub-indicators) ----
# header rows 1..9 carry the spec; row 10 = column titles; councils from row 11.
entry = wb.create_sheet('Advanced Entry'); entry.sheet_view.showGridLines = False
meta_rows = ['Indicator', 'Owner', 'Component', 'Unit', 'Transform', 'Ref Min', 'Ref Max', 'Direction', 'Weight']
HEADR = 10
for r, lbl in enumerate(meta_rows, 1):
    entry.cell(r, 3, lbl).font = Font(size=9, italic=True, color='64748B')
sc = wb.create_sheet('Advanced 0-10'); sc.sheet_view.showGridLines = False
comp_sheet = wb.create_sheet('Components'); comp_sheet.sheet_view.showGridLines = False

for j, ttl in enumerate(['Region', 'Council', 'Code'], 1):
    for sh in (entry, sc, comp_sheet):
        c = sh.cell(HEADR, j, ttl); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = HEAD

# place each sub-indicator in a column; remember each component's column span (contiguous)
span = {}  # component -> [first_col, last_col]
for ci, a in enumerate(adv):
    col = 4 + ci; L = get_column_letter(col); comp = a['component']
    sg = 'Decrease Risk' if str(a['sign']).startswith('Dec') else 'Increase Risk'
    meta = [a['name'], a['sector'], comp, a['unit'], a['transform'],
            float(a['resolved_min']), float(a['resolved_max']), sg, float(a['weight'])]
    for r, v in enumerate(meta, 1):
        entry.cell(r, col, v).font = Font(size=9, bold=(r == 2), color=('B45309' if r == 2 else '000000'))
    for sh in (entry, sc):
        c = sh.cell(HEADR, col, a['name']); c.font = Font(bold=True, color='FFFFFF', size=8); c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical='top'); sh.column_dimensions[L].width = 12
    s = span.setdefault(comp, [col, col]); s[1] = col

# component columns on the Components sheet
ccol = {}
for k, comp in enumerate(comp_order):
    col = 4 + k
    c = comp_sheet.cell(HEADR, col, comp); c.font = Font(bold=True, color='FFFFFF', size=9); c.fill = HEAD
    c.alignment = Alignment(wrap_text=True, vertical='top'); comp_sheet.column_dimensions[get_column_letter(col)].width = 13
    ccol[comp] = col
natcol = 4 + len(comp_order)
c = comp_sheet.cell(HEADR, natcol, 'Natural (advanced)'); c.font = Font(bold=True, color='FFFFFF', size=9); c.fill = HEAD
c.alignment = Alignment(wrap_text=True, vertical='top'); comp_sheet.column_dimensions[get_column_letter(natcol)].width = 14

for ri, (reg, name, code) in enumerate(councils):
    r = HEADR + 1 + ri
    for sh in (entry, sc, comp_sheet):
        sh.cell(r, 1, reg); sh.cell(r, 2, name); sh.cell(r, 3, code)
    # entry (unlocked, institution-coloured) + standardiser
    for ci, a in enumerate(adv):
        col = 4 + ci; L = get_column_letter(col)
        ec = entry.cell(r, col); ec.fill = inst_fill(a['sector']); ec.protection = Protection(locked=False)
        val = f"'Advanced Entry'!{L}{r}"; tf = f"'Advanced Entry'!{L}$5"; mn = f"'Advanced Entry'!{L}$6"
        mx = f"'Advanced Entry'!{L}$7"; sg = f"'Advanced Entry'!{L}$8"
        x = f'IF({tf}="Logarithm",LN(0.001+{val}),IF({tf}="Exponential",EXP({val}),{val}))'
        base = f"10*(({x})-{mn})/({mx}-{mn})"
        expr = f'IF({sg}="Decrease Risk",10-{base},{base})'
        sc.cell(r, col, f'=IF({val}="","",IFERROR(MAX(0,MIN(10,ROUND({expr},1))),"No data"))').font = Font(size=9, color='1F6F3D')
    # components = weighted average over PRESENT cells only (blank skips, no bias)
    for comp, (c0, c1) in span.items():
        a0, a1 = get_column_letter(c0), get_column_letter(c1)
        rng = f"'Advanced 0-10'!{a0}{r}:{a1}{r}"
        wts = f"'Advanced Entry'!{a0}$9:{a1}$9"
        denom = f"SUMPRODUCT(ISNUMBER({rng})*{wts})"
        numer = f"SUMPRODUCT(IFERROR({rng}*1,0)*{wts})"
        comp_sheet.cell(r, ccol[comp], f'=IF({denom}=0,"",ROUND({numer}/{denom},1))').font = Font(size=9, bold=True, color='B45309')
    # Natural category (advanced) = average over the components present (skips empty components)
    n0, n1 = get_column_letter(4), get_column_letter(3 + len(comp_order))
    crng = f"{n0}{r}:{n1}{r}"
    nd = f"SUMPRODUCT(--ISNUMBER({crng}))"
    nn = f"SUMPRODUCT(IFERROR({crng}*1,0))"
    comp_sheet.cell(r, natcol, f'=IF({nd}=0,"",ROUND({nn}/{nd},1))').font = Font(size=9, bold=True, color='1F3A5F')

for sh in (entry, sc, comp_sheet):
    sh.freeze_panes = 'D11'
    sh.column_dimensions['A'].width = 16; sh.column_dimensions['B'].width = 22; sh.column_dimensions['C'].width = 8
    sh.row_dimensions[HEADR].height = 54
# lock the formula sheets but allow editing the unlocked (coloured) entry cells
entry.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False)
sc.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False)
comp_sheet.protection = SheetProtection(sheet=True, selectLockedCells=False, selectUnlockedCells=False)

out = os.path.join(ROOT, 'public/INFORM_TZ_Advanced_Tool.xlsx')
wb.save(out)
from collections import Counter
print('wrote', out)
print('councils:', len(councils), '| exploded sub-indicators:', len(adv), '| components:', len(comp_order))
print('components:', ', '.join(comp_order))
print('by institution:')
for k, v in Counter(a['sector'] for a in adv).most_common():
    print('  %-8s %d' % (k, v))
